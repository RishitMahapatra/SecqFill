"""
Questionnaire upload + inspection endpoints.

The upload flow is the first place where parsing (app/parsing.py) and
matching (app/routers/matching.py::match_single_question) actually meet:
upload a file → extract questions → match every question → store one
questionnaire_items row per question with its match + confidence bucket,
so a human can review them afterwards.
"""
import logging
import os
import tempfile
import time
from pathlib import Path
from uuid import UUID, uuid4

import openpyxl
from docx import Document
from docx.shared import Inches
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask

from app.database import get_db
from app.models import (
    QuestionnaireItemOut,
    QuestionnaireItemUpdate,
    QuestionnaireUploadResponse,
)
from app.parsing import (
    extract_questions_from_docx,
    extract_questions_from_xlsx,
)
from app.routers.matching import match_single_question

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/companies/{company_id}/questionnaires", tags=["questionnaires"]
)

SUPPORTED_EXTENSIONS = {".xlsx", ".docx"}

# backend/app/routers/questionnaires.py -> backend/
BACKEND_ROOT = Path(__file__).resolve().parent.parent.parent
QUESTIONNAIRE_STORAGE_DIR = BACKEND_ROOT / "storage" / "questionnaires"
QUESTIONNAIRE_STORAGE_DIR.mkdir(parents=True, exist_ok=True)

EXPORT_COLUMN_TITLE = "SecQFill Answer"
EXPORT_CONTENT_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


@router.post("", response_model=QuestionnaireUploadResponse, status_code=201)
def upload_questionnaire(company_id: UUID, file: UploadFile = File(...)):
    filename = file.filename or ""
    _, ext = os.path.splitext(filename.lower())
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            400,
            f"Unsupported file type '{ext}'. Supported: .xlsx, .docx",
        )

    # Generated up front (rather than letting Postgres default it) so the
    # file can be named after the questionnaire's id before the row exists.
    questionnaire_id = uuid4()
    stored_path = QUESTIONNAIRE_STORAGE_DIR / f"{questionnaire_id}{ext}"
    stored_path.write_bytes(file.file.read())
    # Relative to BACKEND_ROOT, not absolute — an absolute path would break
    # the moment this runs on a different machine or deploy path.
    relative_storage_path = str(stored_path.relative_to(BACKEND_ROOT))

    # The file now persists (needed for export later), so parsing reads
    # directly from the saved copy instead of a separate temp file.
    if ext == ".xlsx":
        questions = extract_questions_from_xlsx(str(stored_path))
    else:
        questions = extract_questions_from_docx(str(stored_path))

    with get_db() as conn:
        conn.execute(
            """
            INSERT INTO questionnaires
                (id, company_id, original_filename, file_type, status, storage_path)
            VALUES (%s, %s, %s, %s, 'processing', %s)
            """,
            (
                str(questionnaire_id),
                str(company_id),
                filename,
                ext.lstrip("."),
                relative_storage_path,
            ),
        )

        counts = {"green": 0, "yellow": 0, "red": 0}
        for q in questions:
            match = match_single_question(company_id, q.question_text, conn)
            counts[match.confidence] += 1

            conn.execute(
                """
                INSERT INTO questionnaire_items
                    (questionnaire_id, row_number, question_text,
                     matched_answer_id, confidence_score, status,
                     final_answer_text)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    str(questionnaire_id),
                    q.row_number,
                    q.question_text,
                    str(match.id) if match.id is not None else None,
                    round(match.similarity, 3)
                    if match.similarity is not None
                    else None,
                    match.confidence,
                    match.answer_text,
                ),
            )

        conn.execute(
            "UPDATE questionnaires SET status = 'ready_for_review' WHERE id = %s",
            (str(questionnaire_id),),
        )

    return QuestionnaireUploadResponse(
        id=questionnaire_id,
        status="ready_for_review",
        total_questions=len(questions),
        green_count=counts["green"],
        yellow_count=counts["yellow"],
        red_count=counts["red"],
    )


@router.get(
    "/{questionnaire_id}/items", response_model=list[QuestionnaireItemOut]
)
def list_questionnaire_items(company_id: UUID, questionnaire_id: UUID):
    with get_db() as conn:
        # Verify the questionnaire actually belongs to this company before
        # returning its items — without this check, anyone could read any
        # questionnaire's items by guessing the id.
        exists = conn.execute(
            "SELECT 1 FROM questionnaires WHERE id = %s AND company_id = %s",
            (str(questionnaire_id), str(company_id)),
        ).fetchone()
        if exists is None:
            raise HTTPException(404, "Questionnaire not found")

        rows = conn.execute(
            """
            SELECT id, questionnaire_id, row_number, question_text,
                   matched_answer_id, confidence_score, status,
                   final_answer_text, approved, created_at
            FROM questionnaire_items
            WHERE questionnaire_id = %s
            ORDER BY row_number
            """,
            (str(questionnaire_id),),
        ).fetchall()

    return [_row_to_item(r) for r in rows]


@router.patch(
    "/{questionnaire_id}/items/{item_id}", response_model=QuestionnaireItemOut
)
def update_questionnaire_item(
    company_id: UUID,
    questionnaire_id: UUID,
    item_id: UUID,
    payload: QuestionnaireItemUpdate,
):
    fields, values = [], []
    if payload.final_answer_text is not None:
        fields.append("final_answer_text = %s")
        values.append(payload.final_answer_text)
    if payload.approved is not None:
        fields.append("approved = %s")
        values.append(payload.approved)
    if not fields:
        raise HTTPException(400, "No fields to update")

    with get_db() as conn:
        # Chain the full ownership check: item belongs to questionnaire
        # belongs to company. Without this, a caller could patch any item
        # by guessing its id regardless of which company/questionnaire the
        # URL claims it's under.
        exists = conn.execute(
            """
            SELECT 1
            FROM questionnaire_items qi
            JOIN questionnaires q ON q.id = qi.questionnaire_id
            WHERE qi.id = %s AND qi.questionnaire_id = %s AND q.company_id = %s
            """,
            (str(item_id), str(questionnaire_id), str(company_id)),
        ).fetchone()
        if exists is None:
            raise HTTPException(404, "Questionnaire item not found")

        values.append(str(item_id))
        row = conn.execute(
            f"""
            UPDATE questionnaire_items SET {", ".join(fields)}
            WHERE id = %s
            RETURNING id, questionnaire_id, row_number, question_text,
                      matched_answer_id, confidence_score, status,
                      final_answer_text, approved, created_at
            """,
            values,
        ).fetchone()

    return _row_to_item(row)


@router.post(
    "/{questionnaire_id}/rematch", response_model=QuestionnaireUploadResponse
)
def rematch_questionnaire(
    company_id: UUID, questionnaire_id: UUID, use_llm_judge: bool = False
):
    """
    Re-run matching for every item in an already-uploaded questionnaire.
    Lets a growing answer_bank fix items that were wrong at upload time
    without re-uploading the file. Approved items are left untouched —
    a human already signed off on those, so a fresh match shouldn't
    silently overwrite their decision.

    use_llm_judge defaults to False, matching match_single_question()'s
    own default — the LLM judge hasn't been performance-tested across a
    full questionnaire yet. Passing it explicitly here lets it be timed
    against real data before deciding whether to make it the default.
    """
    start_time = time.monotonic()

    with get_db() as conn:
        questionnaire_row = conn.execute(
            "SELECT status FROM questionnaires WHERE id = %s AND company_id = %s",
            (str(questionnaire_id), str(company_id)),
        ).fetchone()
        if questionnaire_row is None:
            raise HTTPException(404, "Questionnaire not found")
        questionnaire_status = questionnaire_row[0]

        items = conn.execute(
            """
            SELECT id, question_text, approved, status
            FROM questionnaire_items
            WHERE questionnaire_id = %s
            ORDER BY row_number
            """,
            (str(questionnaire_id),),
        ).fetchall()

        counts = {"green": 0, "yellow": 0, "red": 0}
        for item_id, question_text, is_approved, current_status in items:
            if is_approved:
                if current_status in counts:
                    counts[current_status] += 1
                continue

            match = match_single_question(
                company_id, question_text, conn, use_llm_judge=use_llm_judge
            )
            counts[match.confidence] += 1

            conn.execute(
                """
                UPDATE questionnaire_items
                SET matched_answer_id = %s,
                    confidence_score = %s,
                    status = %s,
                    final_answer_text = %s
                WHERE id = %s
                """,
                (
                    str(match.id) if match.id is not None else None,
                    round(match.similarity, 3)
                    if match.similarity is not None
                    else None,
                    match.confidence,
                    match.answer_text,
                    str(item_id),
                ),
            )

    elapsed = time.monotonic() - start_time
    logger.warning(
        "Rematch completed: %d items in %.1fs, use_llm_judge=%s",
        len(items),
        elapsed,
        use_llm_judge,
    )

    return QuestionnaireUploadResponse(
        id=questionnaire_id,
        status=questionnaire_status,
        total_questions=len(items),
        green_count=counts["green"],
        yellow_count=counts["yellow"],
        red_count=counts["red"],
    )


@router.get("/{questionnaire_id}/export")
def export_questionnaire(company_id: UUID, questionnaire_id: UUID):
    """
    Regenerate the original file with a "SecQFill Answer" column appended,
    filled from each item's final_answer_text and aligned by row_number.
    Every export is a fresh rebuild from current item data — no diffing
    against a previous export, no detection of an existing answer column.
    """
    with get_db() as conn:
        questionnaire_row = conn.execute(
            """
            SELECT original_filename, file_type, storage_path
            FROM questionnaires
            WHERE id = %s AND company_id = %s
            """,
            (str(questionnaire_id), str(company_id)),
        ).fetchone()
        if questionnaire_row is None:
            raise HTTPException(404, "Questionnaire not found")

        original_filename, file_type, storage_path = questionnaire_row

        if not storage_path:
            raise HTTPException(
                404,
                "No stored file for this questionnaire — it was uploaded "
                "before file persistence was added and can't be exported.",
            )

        source_path = BACKEND_ROOT / storage_path
        if not source_path.is_file():
            raise HTTPException(404, "Stored file is missing on disk")

        items = conn.execute(
            """
            SELECT row_number, final_answer_text
            FROM questionnaire_items
            WHERE questionnaire_id = %s
            ORDER BY row_number
            """,
            (str(questionnaire_id),),
        ).fetchall()

    if file_type not in EXPORT_CONTENT_TYPES:
        raise HTTPException(400, f"Unsupported file type '{file_type}' for export")

    export_ext = f".{file_type}"
    with tempfile.NamedTemporaryFile(suffix=export_ext, delete=False) as tmp:
        export_path = Path(tmp.name)

    if file_type == "xlsx":
        _write_xlsx_export(source_path, items, export_path)
    else:
        _write_docx_export(source_path, items, export_path)

    stem, dot_ext = os.path.splitext(original_filename)
    download_name = f"{stem}_completed{dot_ext}"

    return FileResponse(
        path=export_path,
        media_type=EXPORT_CONTENT_TYPES[file_type],
        filename=download_name,
        background=BackgroundTask(export_path.unlink, missing_ok=True),
    )


def _write_xlsx_export(source_path: Path, items, export_path: Path) -> None:
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active
    answer_col = ws.max_column + 1
    ws.cell(row=1, column=answer_col, value=EXPORT_COLUMN_TITLE)
    for row_number, final_answer_text in items:
        ws.cell(row=row_number, column=answer_col, value=final_answer_text or "")
    wb.save(export_path)


def _write_docx_export(source_path: Path, items, export_path: Path) -> None:
    doc = Document(source_path)
    table = doc.tables[0]
    answer_col = table.add_column(width=Inches(2))
    answer_col.cells[0].text = EXPORT_COLUMN_TITLE
    for row_number, final_answer_text in items:
        row_idx = row_number - 1
        if 0 <= row_idx < len(answer_col.cells):
            answer_col.cells[row_idx].text = final_answer_text or ""
    doc.save(export_path)


def _row_to_item(row) -> QuestionnaireItemOut:
    return QuestionnaireItemOut(
        id=row[0],
        questionnaire_id=row[1],
        row_number=row[2],
        question_text=row[3],
        matched_answer_id=row[4],
        confidence_score=float(row[5]) if row[5] is not None else None,
        status=row[6],
        final_answer_text=row[7],
        approved=row[8],
        created_at=row[9],
    )
